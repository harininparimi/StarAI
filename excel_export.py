from io import BytesIO
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def dataframe_to_excel(dataframe: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="StarAI Leads")
        ws = writer.book["StarAI Leads"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        for idx, column in enumerate(dataframe.columns, start=1):
            values = [str(column)] + [
                "" if v is None else str(v)
                for v in dataframe[column].tolist()
            ]
            ws.column_dimensions[get_column_letter(idx)].width = min(
                max(len(v) for v in values) + 2, 55
            )

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    return output.getvalue()
